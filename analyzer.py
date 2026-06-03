import os
import csv
import numpy as np
import sounddevice as sd
import soundfile as sf
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

from scipy.signal import (
    find_peaks,
    hilbert,
    butter,
    filtfilt
)
from scipy.optimize import curve_fit
from scipy.fft import rfft, rfftfreq

FS = 44100
DURATION = 3

OUTPUT_DIR = "measurements"

CSV_FILE = os.path.join(OUTPUT_DIR, "results.csv")
XLSX_FILE = os.path.join(OUTPUT_DIR, "results.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

if not os.path.exists(CSV_FILE):
    with open(CSV_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "measurement",
            "f1", "f2", "f3", "f4", "f5",
            "delta",
            "alpha",
            "q",
            "rms",
            "energy",
            "area",
            "f2_f1",
            "f3_f1",
            "band_energy_ratio"
        ])

if not os.path.exists(XLSX_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    ws.append([
        "measurement",
        "f1", "f2", "f3", "f4", "f5",
        "delta",
        "alpha",
        "q",
        "rms",
        "energy",
        "area",
        "f2_f1",
        "f3_f1",
        "band_energy_ratio"
    ])

    wb.save(XLSX_FILE)


def record_signal():
    print("\nУдарьте по образцу...\n")

    signal = sd.rec(
        int(DURATION * FS),
        samplerate=FS,
        channels=1,
        dtype="float64"
    )

    sd.wait()

    return signal.flatten()


def save_wav(signal):
    files = len(os.listdir(OUTPUT_DIR))

    filename = os.path.join(
        OUTPUT_DIR,
        f"measurement_{files+1:03d}.wav"
    )

    sf.write(filename, signal, FS)

    print(f"Файл сохранен: {filename}")

    return filename


def extract_impact_region(signal):
    abs_signal = np.abs(signal)

    impact_idx = np.argmax(abs_signal)

    start = impact_idx
    end = min(len(signal), start + 2048)

    segment = signal[start:end]

    # если вдруг меньше 4096 отсчетов — дополняем нулями
    if len(segment) < 2048:
        segment = np.pad(segment, (0, 2048 - len(segment)))

    return segment


def analyze_fft(signal):

    N = len(signal)

    signal = signal - np.mean(signal)
    signal = signal / (np.max(np.abs(signal)) + 1e-12)

    spectrum = np.abs(rfft(signal))
    freqs = rfftfreq(N, 1 / FS)

    valid = freqs > 100

    peaks, _ = find_peaks(
        spectrum[valid],
        height=np.max(spectrum[valid]) * 0.1,
        distance=20
    )

    peak_freqs = freqs[valid][peaks]
    peak_amps = spectrum[valid][peaks]

    order = np.argsort(peak_amps)[::-1]

    peak_freqs = peak_freqs[order]
    peak_amps = peak_amps[order]

    return freqs, spectrum, peak_freqs[:10], peak_amps[:10]


# --- Дополнительные функции ---

def bandpass_filter(signal, lowcut, highcut, fs, order=4):
    nyquist = 0.5 * fs

    signal = signal - np.mean(signal)

    lowcut = max(20.0, lowcut)
    highcut = min(highcut, nyquist * 0.9)

    low = lowcut / nyquist
    high = highcut / nyquist

    if low >= high:
        return signal.copy()

    b, a = butter(order, [low, high], btype="band")

    filtered = filtfilt(b, a, signal)

    filtered = np.nan_to_num(filtered)

    max_val = np.max(np.abs(filtered))

    if max_val > 0:
        filtered = filtered / max_val

    return filtered


def get_envelope(signal):
    analytic = hilbert(signal)
    return np.abs(analytic)


def exp_decay(t, A0, alpha):
    return A0 * np.exp(-alpha * t)


def calculate_alpha(envelope):
    if len(envelope) == 0:
        return np.nan, np.nan

    max_env = np.max(envelope)

    if max_env <= 0:
        return np.nan, np.nan

    t = np.arange(len(envelope)) / FS

    valid = envelope > max_env * 0.05

    t_fit = t[valid]
    env_fit = envelope[valid]

    if len(env_fit) < 10:
        return np.nan, np.nan

    try:
        popt, _ = curve_fit(
            exp_decay,
            t_fit,
            env_fit,
            p0=[max_env, 1],
            maxfev=10000
        )

        A0, alpha = popt

        return alpha, A0

    except Exception:
        return np.nan, np.nan


def calculate_log_decrement(envelope):
    peaks, _ = find_peaks(
        envelope,
        height=np.max(envelope) * 0.1,
        distance=max(20, int(FS / 500))
    )

    if len(peaks) < 2:
        return np.nan

    amps = envelope[peaks]
    amps = amps[amps > 0]

    if len(amps) < 2:
        return np.nan

    return np.log(amps[0] / amps[-1]) / max(1, len(amps) - 1)


def calculate_energy(signal):
    signal = np.nan_to_num(signal)
    return float(np.sum(signal * signal))


def calculate_area(envelope):
    envelope = np.nan_to_num(envelope)
    return float(np.trapezoid(envelope, dx=1 / FS))


def calculate_rms(signal):
    signal = np.nan_to_num(signal)
    return float(np.sqrt(np.mean(signal * signal)))


def calculate_q(delta):
    if np.isnan(delta) or delta <= 0:
        return np.nan

    return np.pi / (delta + 1e-12)


def calculate_peak_frequency_ratios(peak_freqs):
    if len(peak_freqs) < 2:
        return np.nan, np.nan

    f1 = peak_freqs[0]
    f2 = peak_freqs[1]

    ratio_21 = f2 / f1 if f1 > 0 else np.nan

    if len(peak_freqs) >= 3:
        ratio_31 = peak_freqs[2] / f1 if f1 > 0 else np.nan
    else:
        ratio_31 = np.nan

    return ratio_21, ratio_31


def calculate_band_energy_ratio(freqs, spectrum, f0):
    total_energy = np.sum(spectrum ** 2)

    if total_energy <= 0:
        return np.nan

    mask = (freqs >= f0 - 50) & (freqs <= f0 + 50)

    band_energy = np.sum(spectrum[mask] ** 2)

    return float(band_energy / total_energy)


def save_results(peak_freqs,
                 delta,
                 alpha,
                 q,
                 rms,
                 energy,
                 area,
                 ratio_21,
                 ratio_31,
                 band_ratio):

    measurement_id = len(os.listdir(OUTPUT_DIR))

    freqs = list(peak_freqs[:5])

    while len(freqs) < 5:
        freqs.append(np.nan)

    with open(CSV_FILE, "a", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            measurement_id,
            freqs[0],
            freqs[1],
            freqs[2],
            freqs[3],
            freqs[4],
            delta,
            alpha,
            q,
            rms,
            energy,
            area,
            ratio_21,
            ratio_31,
            band_ratio
        ])

    # --- SAVE TO EXCEL ---
    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append([
            "measurement",
            "f1", "f2", "f3", "f4", "f5",
            "delta",
            "alpha",
            "q",
            "rms",
            "energy",
            "area",
            "f2_f1",
            "f3_f1",
            "band_energy_ratio"
        ])

    ws.append([
        measurement_id,
        freqs[0],
        freqs[1],
        freqs[2],
        freqs[3],
        freqs[4],
        delta,
        alpha,
        q,
        rms,
        energy,
        area,
        ratio_21,
        ratio_31,
        band_ratio
    ])

    wb.save(XLSX_FILE)


def plot_results(signal,
                 freqs,
                 spectrum,
                 peak_freqs,
                 signal_cut,
                 envelope,
                 f0):

    t_signal = np.arange(len(signal)) / FS
    t_mode = np.arange(len(signal_cut)) / FS

    fig, ax = plt.subplots(3, 1, figsize=(12, 8), constrained_layout=True)

    ax[0].plot(t_signal, signal)
    ax[0].set_title("Сигнал")
    ax[0].grid(True)

    ax[1].plot(freqs, spectrum)

    for f in peak_freqs[:5]:
        ax[1].axvline(f, linestyle="--")

    ax[1].set_xlim(0, min(FS/2, f0 * 3))
    ax[1].set_title("FFT")
    ax[1].grid(True)

    ax[2].plot(t_mode, signal_cut, label="Сигнал")
    ax[2].plot(t_mode, envelope, linewidth=2, label="Огибающая")
    ax[2].legend()
    ax[2].set_title("Выделенная мода")
    ax[2].grid(True)

    plt.show()


def main():

    signal = record_signal()

    save_wav(signal)

    signal = extract_impact_region(signal)

    freqs, spectrum, peak_freqs, peak_amps = analyze_fft(signal)

    print("\nНайденные пики:")

    for i, f in enumerate(peak_freqs[:10], start=1):
        print(f"{i}. {f:.2f} Гц")

    if len(peak_freqs) > 0:
        print(
            f"\nОсновная частота: "
            f"{peak_freqs[0]:.2f} Гц"
        )

    f0 = peak_freqs[0]

    bandwidth = max(50, f0 * 0.3)

    signal_filt = bandpass_filter(
        signal,
        max(10, f0 - bandwidth),
        f0 + bandwidth,
        FS
    )

    # --- ALIGNMENT: start mode from impact peak ---
    impact_idx = np.argmax(np.abs(signal_filt))
    signal_filt = np.roll(signal_filt, -impact_idx)

    if f0 < 120:
        print("ВНИМАНИЕ: обнаружена очень низкая частота, возможен захват шума или резонанса стола.")

    envelope = get_envelope(signal_filt)
    envelope = np.nan_to_num(envelope)

    cutoff = int(len(envelope) * 0.95)
    envelope = envelope[:cutoff]
    signal_filt_cut = signal_filt[:cutoff]

    print(f"Максимум огибающей: {np.max(envelope):.6f}")

    alpha, _ = calculate_alpha(envelope)

    delta = calculate_log_decrement(envelope)

    q = calculate_q(delta)

    energy = calculate_energy(signal_filt_cut)

    rms = calculate_rms(signal_filt_cut)

    area = calculate_area(envelope)

    ratio_21, ratio_31 = calculate_peak_frequency_ratios(peak_freqs)

    band_ratio = calculate_band_energy_ratio(
        freqs,
        spectrum,
        f0
    )

    print("\n========== РЕЗУЛЬТАТЫ ==========")
    print(f"Основная частота: {f0:.2f} Гц")
    print(f"Логарифмический декремент: {delta:.6f}")
    print(f"Коэффициент затухания α: {alpha:.6f}")
    print(f"Добротность Q: {q:.2f}")
    print(f"RMS: {rms:.6f}")
    print(f"Энергия: {energy:.2f}")
    print(f"Площадь огибающей: {area:.2f}")
    print(f"f2/f1: {ratio_21:.4f}")
    print(f"f3/f1: {ratio_31:.4f}")
    print(f"Доля энергии основной моды: {band_ratio:.4f}")

    print("\nПервые 5 мод:")
    for i, freq in enumerate(peak_freqs[:5], start=1):
        print(f"f{i} = {freq:.2f} Гц")

    save_results(
        peak_freqs,
        delta,
        alpha,
        q,
        rms,
        energy,
        area,
        ratio_21,
        ratio_31,
        band_ratio
    )

    print(f"\nРезультаты сохранены в: {CSV_FILE}")

    plot_results(
        signal,
        freqs,
        spectrum,
        peak_freqs,
        signal_filt_cut,
        envelope,
        f0
    )


if __name__ == "__main__":
    main()